'''生成Excel文件'''
import openpyxl

def new_excel(path):
    '''新建excel，改变sheet名字，并写入内容'''
    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    work_sheet.title = 'data'
    # ws.cell(row=1, column=2).value = 1
    work_book.save(path)


def write_pdf_names(name_list, path):
    '''加载PDF文件名列'''
    work_book = openpyxl.load_workbook(path)
    work_sheet_1 = work_book['data']
    num = len(name_list)
    for i in range(num):
        work_sheet_1.cell(row=i + 2, column=1).value = name_list[i]
    work_book.save(path)


def find_location(member, list_to_find):
    '''将两个列表对齐'''
    length = len(list_to_find)
    for i in range(length):
        if member == list_to_find[i]:
            return i
    return length


def write_pdf_data(labels, values, path):
    '''写入PDF数据'''
    work_book = openpyxl.load_workbook(path)
    work_sheet_1 = work_book['data']
    pdf_num = len(labels)
    label0 = labels[0]
    labels_num = len(label0)
    for i in range(labels_num):
        work_sheet_1.cell(row=1, column=i + 2).value = label0[i]
    for i in range(pdf_num):
        single_pdf_labels = labels[i]
        single_pdf_values = values[i]
        for j in range(labels_num):
            location = find_location(single_pdf_labels[j], label0)
            new_str = ''
            if len(single_pdf_values[location]) == 1:
                new_str = single_pdf_values[location][0]
                if isinstance(new_str, str):
                    if new_str.count('/') == 2:
                        new_str_list = new_str.split('/')
                        new_str = new_str_list[2] + '/' + new_str_list[1] + '/' + new_str_list[0]
                work_sheet_1.cell(row=i + 2, column=location + 2).value = new_str
            else:
                for k in range(len(single_pdf_values[location])):
                    if k == 0:
                        new_str = single_pdf_values[location][0]
                    else:
                        new_str = new_str + ';' + single_pdf_values[location][k]
                work_sheet_1.cell(row=i + 2, column=location + 2).value = new_str
    work_book.save(path)
